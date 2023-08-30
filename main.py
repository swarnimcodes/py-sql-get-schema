import json
import pyodbc
import openpyxl
import customtkinter as ctk
import tkinter as tk
from tkinter import filedialog

# Set the appearance mode and color theme for customtkinter
ctk.set_appearance_mode("system")  # Modes: system (default), light, dark
ctk.set_default_color_theme("dark-blue")  # Themes: blue (default), dark-blue, green

def get_database_names():
    return entry_databases.get().split(',')

def get_server():
    return entry_server.get()

def get_username():
    return entry_username.get()

def get_password():
    return entry_password.get()

def perform_schema_comparison(source_database, target_databases):
    # Server Details
    server = get_server()
    username = get_username()
    password = get_password()

    source_schema_info = None
    with open(f'schema_information_{source_database}.json', 'r') as json_file:
        source_schema_info = json.load(json_file)

    # Create an Excel workbook
    workbook = openpyxl.Workbook()

    for target_database in target_databases:
        sheet = workbook.create_sheet(title=target_database)  # Create a new sheet for each target database

        # Set headers
        sheet["A1"] = "Schema"
        sheet["B1"] = "Table Name"
        sheet["C1"] = "Type of Error"
        sheet["D1"] = "Column Name"
        sheet["E1"] = f"Specification in {source_database}"
        sheet["F1"] = f"Specification in {target_database}"

        row = 2  # Start from the second row

        with open(f'schema_information_{target_database}.json', 'r') as json_file:
            target_schema_info = json.load(json_file)

        for schema in source_schema_info:
            for table_name in source_schema_info[schema]:
                source_columns = source_schema_info[schema][table_name]
                target_columns = target_schema_info[schema].get(table_name, [])

                for col_info_source in source_columns:
                    col_name_source = col_info_source["column_name"]
                    col_info_target = next((col for col in target_columns if col["column_name"] == col_name_source), None)

                    if col_info_target is None:
                        sheet[f"A{row}"] = schema
                        sheet[f"B{row}"] = table_name
                        sheet[f"C{row}"] = "Missing Column"
                        sheet[f"D{row}"] = col_name_source
                        sheet[f"E{row}"] = str(col_info_source)
                        sheet[f"F{row}"] = ""
                        sheet[f"C{row}"].fill = openpyxl.styles.PatternFill(start_color="E6B8B7", end_color="E6B8B7", fill_type="solid")
                        row += 1
                    elif col_info_source != col_info_target:
                        sheet[f"A{row}"] = schema
                        sheet[f"B{row}"] = table_name
                        sheet[f"C{row}"] = "Different Specification"
                        sheet[f"D{row}"] = col_name_source
                        sheet[f"E{row}"] = str(col_info_source)
                        sheet[f"F{row}"] = str(col_info_target)
                        sheet[f"C{row}"].fill = openpyxl.styles.PatternFill(start_color="CCC0DA", end_color="CCC0DA", fill_type="solid")
                        row += 1

        print(f"Comparison report generated for {target_database}")

    # Save the Excel file
    excel_filename = "schema_comparison_report.xlsx"
    workbook.remove(workbook.active)  # Remove the default sheet
    workbook.save(excel_filename)

    print(f"Schema comparison report saved to {excel_filename}")


root = ctk.CTk()
# root.geometry("600x400")
root.after(0, lambda:root.state('zoomed'))
root.title("Database Schema Comparer")

label_server = ctk.CTkLabel(master=root, text="Enter Server Address:")
label_server.place(relx=0.35, rely=0.10, anchor=ctk.CENTER)
entry_server = ctk.CTkEntry(master=root)
entry_server.place(relx=0.75, rely=0.10, anchor=ctk.CENTER)

label_source = ctk.CTkLabel(master=root, text="Enter Source Database Name:")
label_source.place(relx=0.35, rely=0.25, anchor=ctk.CENTER)
entry_source = ctk.CTkEntry(master=root)
entry_source.place(relx=0.75, rely=0.25, anchor=ctk.CENTER)

label_databases = ctk.CTkLabel(master=root, text="Enter Target Database Names (comma-separated):")
label_databases.place(relx=0.35, rely=0.40, anchor=ctk.CENTER)
entry_databases = ctk.CTkEntry(master=root)
entry_databases.place(relx=0.75, rely=0.40, anchor=ctk.CENTER)

label_username = ctk.CTkLabel(master=root, text="Enter your username:")
label_username.place(relx=0.35, rely=0.55, anchor=ctk.CENTER)
entry_username = ctk.CTkEntry(master=root)
entry_username.place(relx=0.75, rely=0.55, anchor=ctk.CENTER)

label_password = ctk.CTkLabel(master=root, text="Enter your password:")
label_password.place(relx=0.35, rely=0.70, anchor=ctk.CENTER)
entry_password = ctk.CTkEntry(master=root, show="*")
entry_password.place(relx=0.75, rely=0.70, anchor=ctk.CENTER)

button = ctk.CTkButton(master=root, text="Submit", command=root.quit)  # Using root.quit to close the GUI
button.place(relx=0.5, rely=0.85, anchor=ctk.CENTER)

root.mainloop()

# After the GUI is closed, the program continues here
source_database = entry_source.get()
target_databases = get_database_names()

perform_schema_comparison(source_database, target_databases)
